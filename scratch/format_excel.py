import openpyxl
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter

def format_excel():
    src_path = "02.07.2026 - Mentorados Atualizada (1).xlsx"
    
    wb = openpyxl.load_workbook(src_path)
    
    original_sheets = ['2026', 'Antigos', 'Acompanhamento - EMPR', 'Acompanhamento - EMPR (2)']
    for s_name in list(wb.sheetnames):
        if s_name not in original_sheets:
            del wb[s_name]
            
    f_header = Font(name="Calibri", size=11, bold=True, color="FFFFFF")
    f_body = Font(name="Calibri", size=11, color="000000")
    
    # Header fill matching Screenshot 2 (Medium Dark Gray #595959)
    fill_header = PatternFill(start_color="595959", end_color="595959", fill_type="solid")
    
    # Status Highlights: GREEN (#C6EFCE fill, #006100 font) matching Screenshot 2 EXACTLY
    fill_ok = PatternFill(start_color="C6EFCE", end_color="C6EFCE", fill_type="solid")
    font_ok = Font(name="Calibri", size=11, color="006100")
    
    # Scheduled dates: soft blue
    fill_info = PatternFill(start_color="EEF4FF", end_color="EEF4FF", fill_type="solid")
    font_info = Font(name="Calibri", size=11, color="3767B1")
    
    # Pendencies / Notes: soft amber
    fill_warn = PatternFill(start_color="FFEB9C", end_color="FFEB9C", fill_type="solid")
    font_warn = Font(name="Calibri", size=11, color="9C6500")
    
    thin_border_side = Side(border_style="thin", color="D9D9D9")
    thin_border = Border(left=thin_border_side, right=thin_border_side, top=thin_border_side, bottom=thin_border_side)

    align_center = Alignment(horizontal="center", vertical="center", wrap_text=False)
    align_left = Alignment(horizontal="left", vertical="center", wrap_text=False)

    for sheet_name in wb.sheetnames:
        ws = wb[sheet_name]
        ws.views.sheetView[0].showGridLines = True
        
        max_r = ws.max_row
        max_c = ws.max_column
        
        if max_r == 0 or max_c == 0:
            continue
            
        for r in range(1, max_r + 1):
            row_vals = [ws.cell(r, c).value for c in range(1, max_c + 1)]
            is_header = (r == 1) or ("Mentorados" in [str(v) for v in row_vals if v]) or ("Sessão 1" in [str(v) for v in row_vals if v])
            
            ws.row_dimensions[r].height = 24 if is_header else 20
                
            for c in range(1, max_c + 1):
                cell = ws.cell(r, c)
                val_str = str(cell.value or "").strip()
                
                # Normalize cell value if it has "✓ OK" back to "ok"
                if val_str in ["✓ OK", "✓ ok", "OK", "Ok"]:
                    cell.value = "ok"
                    val_str = "ok"
                    
                cell.border = thin_border
                
                if is_header:
                    cell.fill = fill_header
                    cell.font = f_header
                    cell.alignment = align_center
                else:
                    cell.font = f_body
                    cell.alignment = align_left if c in [2, 3, 4, 5] else align_center
                    
                    val_lower = val_str.lower()
                    # Check for OK / completed status -> MUST BE GREEN (#C6EFCE fill, #006100 text)!
                    if "ok" in val_lower or val_lower in ["concluído", "concluido", "concluída", "concluida"]:
                        cell.fill = fill_ok
                        cell.font = font_ok
                    elif any(ch in val_str for ch in ["/", "h", ":"]) and len(val_str) > 3:
                        cell.fill = fill_info
                        cell.font = font_info
                    elif val_str and val_lower not in ["none", "null", ""]:
                        if c > 5:
                            cell.fill = fill_warn
                            cell.font = font_warn
                            
        # Dynamic Column Widths
        for c in range(1, max_c + 1):
            col_letter = get_column_letter(c)
            max_len = 10
            for r in range(1, max_r + 1):
                v = str(ws.cell(r, c).value or "")
                if len(v) > max_len and len(v) < 60:
                    max_len = len(v)
            ws.column_dimensions[col_letter].width = max_len + 4
            
        # Freeze Panes & AutoFilter
        ws.freeze_panes = "C2"
        ws.auto_filter.ref = f"A1:{get_column_letter(max_c)}{max_r}"

    saved_files = []
    targets = [
        "02.07.2026 - Mentorados Atualizada (1).xlsx",
        "02.07.2026 - Mentorados Atualizada (Visualmente Aprimorada).xlsx",
        "02.07.2026 - Mentorados Atualizada (Visualmente Aprimorada) Verde.xlsx"
    ]
    for target in targets:
        try:
            wb.save(target)
            saved_files.append(target)
        except PermissionError:
            pass

    print("Formatting complete! OK cells are GREEN:", saved_files)

if __name__ == "__main__":
    format_excel()
