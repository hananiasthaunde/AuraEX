import sys
import openpyxl

sys.stdout.reconfigure(encoding='utf-8')

file1 = '02.07.2026 - Mentorados Atualizada (1).xlsx'
file2 = '02.07.2026 - Mentorados Atualizada (Visualmente Aprimorada).xlsx'

wb1 = openpyxl.load_workbook(file1, data_only=True)
wb2 = openpyxl.load_workbook(file2, data_only=True)

print("=== AUDITORIA E COMPARAÇÃO DE DADOS ===")
print("Abas no Ficheiro 1:", wb1.sheetnames)
print("Abas no Ficheiro 2:", wb2.sheetnames)

target_sheets = ['2026', 'Antigos', 'Acompanhamento - EMPR', 'Acompanhamento - EMPR (2)']
total_cells_checked = 0
differences = []

for s_name in target_sheets:
    if s_name not in wb1.sheetnames:
        differences.append(f"Aba '{s_name}' não encontrada no arquivo 1.")
        continue
    if s_name not in wb2.sheetnames:
        differences.append(f"Aba '{s_name}' não encontrada no arquivo 2.")
        continue

    ws1 = wb1[s_name]
    ws2 = wb2[s_name]

    max_r = max(ws1.max_row, ws2.max_row)
    max_c = max(ws1.max_column, ws2.max_column)

    sheet_cells = 0
    sheet_diffs = 0

    for r in range(1, max_r + 1):
        for c in range(1, max_c + 1):
            sheet_cells += 1
            total_cells_checked += 1
            val1 = ws1.cell(r, c).value
            val2 = ws2.cell(r, c).value

            norm1 = str(val1).strip() if val1 is not None else ''
            norm2 = str(val2).strip() if val2 is not None else ''

            if norm1 != norm2:
                if norm1 == '' and norm2 == '':
                    continue
                differences.append(f"[{s_name}] Linha {r}, Coluna {c}: File1='{val1}' | File2='{val2}'")
                sheet_diffs += 1

    print(f"✓ Aba '{s_name}': {sheet_cells} células verificadas. Diferenças: {sheet_diffs}")

print(f"\nTotal de células auditadas nas 4 abas principais: {total_cells_checked}")

if not differences:
    print("\n✅ RESULTADO: TODOS OS DADOS, NOMES, E-MAILS, TELEFONES, SESSÕES E OBSERVAÇÕES SÃO 100% IGUAIS!")
else:
    print(f"\n⚠️ Diferenças encontradas ({len(differences)}):")
    for d in differences:
        print(" -", d)
